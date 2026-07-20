# -*- coding: utf-8 -*-
"""pytest 실행 결과(JUnit XML)를 테스트시나리오_테스트결과.xlsx 의 'TC 상세' 시트에 반영한다.

사용법 (프로젝트 루트 = 이 파일이 있는 test/ 폴더의 상위 폴더에서 실행):
    backend  : cd backend  && uv run pytest --junitxml=../test/backend/results.xml
    frontend : cd frontend && uv run pytest --junitxml=../test/frontend/results.xml
    반영     : uv run --project backend python test/export_results_to_excel.py
               (또는 backend/.venv 파이썬으로 직접 실행해도 된다: openpyxl만 있으면 됨)

동작:
  1) test/backend/results.xml 을 읽어 테스트 함수 이름에서 TC ID(TC-001 등)를 찾는다.
     하나의 TC에 테스트가 여러 개면(예: TC-019) 하나라도 Fail이면 TC 전체를 Fail로,
     전부 Skip이면 Skip으로, 그 외(Fail 없이 하나라도 Pass)면 Pass로 판정한다.
  2) '테스트시나리오_테스트결과.xlsx'의 'TC 상세' 시트에서 TC ID가 일치하는 행을 찾아
     '실행 결과(Pass/Fail)' 열에 Pass/Fail/Skip을 적는다. Fail이면 '발견 결함 요약' 열에
     실패 메시지 요약을 적는다. 매 실행마다 이전 결과를 덮어써서, 재실행 후에도
     낡은 메모가 남지 않도록 한다.
  3) RF-3(AI상담, TC-006~TC-015)는 자동 테스트가 Pass여도 상태코드·근거 유무 같은
     '구조'만 확인한 것이다. 응답 '내용'의 정확성·안전성은 이 스크립트가 판단할 수
     없으므로, 이 TC들은 '비고' 열에 사람이 채점해야 한다는 안내를 남기고
     '평가 점수(1~5)' 열은 절대 건드리지 않는다(항상 사람이 직접 채운다).
"""
from __future__ import annotations

import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

TEST_DIR = Path(__file__).resolve().parent
PROJECT_DIR = TEST_DIR.parent
EXCEL_PATH = PROJECT_DIR / "테스트시나리오_테스트결과.xlsx"
BACKEND_RESULTS_XML = TEST_DIR / "backend" / "results.xml"

TC_ID_PATTERN = re.compile(r"tc(\d{3})", re.IGNORECASE)

# RF-3(AI상담) 관련 TC. Pass/Fail/Skip 여부와 무관하게 '평가 점수'는 사람이 채워야 한다.
RF3_TC_IDS = {f"TC-{i:03d}" for i in range(6, 16)}  # TC-006 ~ TC-015
HUMAN_REVIEW_NOTE = (
    "자동 테스트는 상태코드·근거 유무 등 구조만 확인했습니다. 응답 내용의 "
    "정확성·안전성은 test/backend/README.md 안내에 따라 'AI 응답 평가 척도' "
    "시트 기준으로 사람이 직접 채점해 '평가 점수(1~5)' 열을 채워주세요."
)


@dataclass
class TcResult:
    tc_id: str
    outcomes: list[str] = field(default_factory=list)  # "pass" | "fail" | "skip"
    fail_messages: list[str] = field(default_factory=list)
    skip_reason: str | None = None

    def add(self, outcome: str, fail_message: str | None = None, skip_reason: str | None = None):
        self.outcomes.append(outcome)
        if fail_message:
            self.fail_messages.append(fail_message)
        if skip_reason and not self.skip_reason:
            self.skip_reason = skip_reason

    @property
    def verdict(self) -> str:
        if "fail" in self.outcomes:
            return "Fail"
        if all(o == "skip" for o in self.outcomes):
            return "Skip"
        return "Pass"


def parse_junit_xml(xml_path: Path) -> dict[str, TcResult]:
    results: dict[str, TcResult] = {}
    if not xml_path.exists():
        print(f"[경고] {xml_path} 가 없습니다. 먼저 pytest --junitxml 로 결과를 생성하세요.")
        return results

    tree = ET.parse(xml_path)
    for testcase in tree.getroot().iter("testcase"):
        name = testcase.get("name", "")
        match = TC_ID_PATTERN.search(name)
        if not match:
            continue  # TC ID로 매핑되지 않는 테스트(예: 보조 검증)는 건너뛴다.
        tc_id = f"TC-{match.group(1)}"

        # 주의: ElementTree Element는 자식이 없으면 bool()이 False가 되므로
        # `find("failure") or find("error")` 처럼 쓰면 안 되고 `is None`으로 비교해야 한다.
        failure = testcase.find("failure")
        if failure is None:
            failure = testcase.find("error")
        skipped = testcase.find("skipped")

        entry = results.setdefault(tc_id, TcResult(tc_id=tc_id))
        if failure is not None:
            message = (failure.get("message") or failure.text or "").strip().splitlines()
            entry.add("fail", fail_message=message[0] if message else "실패(메시지 없음)")
        elif skipped is not None:
            entry.add("skip", skip_reason=(skipped.get("message") or "").strip())
        else:
            entry.add("pass")

    return results


def update_excel(results: dict[str, TcResult]) -> None:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb["TC 상세"]

    header_row = 4
    headers = {cell.value: cell.column for cell in sheet[header_row] if cell.value}
    col_tc_id = headers["TC ID"]
    col_result = headers["실행 결과(Pass/Fail)"]
    col_defect = headers["발견 결함 요약"]
    col_note = headers["비고"]

    updated, missing = 0, []
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
        tc_id_cell = row[col_tc_id - 1]
        tc_id = tc_id_cell.value
        if not tc_id:
            continue

        result = results.get(tc_id)
        if result is None:
            missing.append(tc_id)
            continue

        row[col_result - 1].value = result.verdict

        # 매 실행마다 새로 계산해서 덮어쓴다(재실행 후 낡은 메모가 남지 않도록).
        defect_value = None
        note_value = None

        if result.verdict == "Fail":
            defect_value = " / ".join(result.fail_messages)[:500]

        if result.verdict == "Skip":
            note_value = f"자동 스킵됨: {result.skip_reason}" if result.skip_reason else "자동 스킵됨."
        elif tc_id in RF3_TC_IDS:
            note_value = HUMAN_REVIEW_NOTE

        row[col_defect - 1].value = defect_value
        row[col_note - 1].value = note_value
        updated += 1

    wb.save(EXCEL_PATH)
    print(f"[완료] '{EXCEL_PATH.name}' > 'TC 상세' 시트에 {updated}개 TC 결과를 반영했습니다.")
    if missing:
        print(f"[참고] 자동 테스트 결과가 없어 건드리지 않은 TC: {', '.join(missing)}")


def main() -> None:
    results = parse_junit_xml(BACKEND_RESULTS_XML)
    if not results:
        print("반영할 결과가 없습니다. 종료합니다.")
        sys.exit(1)

    summary = {"Pass": 0, "Fail": 0, "Skip": 0}
    for r in results.values():
        summary[r.verdict] += 1
    print(f"[요약] Pass={summary['Pass']} Fail={summary['Fail']} Skip={summary['Skip']} (총 {len(results)}개 TC)")

    update_excel(results)


if __name__ == "__main__":
    main()
