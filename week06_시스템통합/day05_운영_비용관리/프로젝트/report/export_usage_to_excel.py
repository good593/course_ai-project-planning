# -*- coding: utf-8 -*-
"""llm_usage_log(토큰 사용량/비용 기록)을 엑셀 리포트로 뽑아낸다.

day04의 test/export_results_to_excel.py가 pytest 결과를 기존 엑셀 시트에
반영했던 것과 달리, 이 스크립트는 매번 새 워크북을 만든다(누적 집계 대상이라
'덮어쓸 기존 서식'이 없기 때문). 대신 강의자료 '운영 지표 예시' 표와 동일한
두 시트를 만든다.
  - '일별 요약' : 날짜별 호출 수·토큰·비용·평균 응답시간·실패율
  - '호출 상세' : 원자료(디버깅/근거 확인용, 최신순)

사용법 (day05 프로젝트 루트에서):
    uv run --project backend python report/export_usage_to_excel.py --days 30
    (backend/.venv 파이썬으로 직접 실행해도 된다: SQLAlchemy + openpyxl만 있으면 됨)
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

REPORT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = REPORT_DIR.parent
BACKEND_DIR = PROJECT_DIR / "backend"
EXCEL_PATH = REPORT_DIR / "토큰_비용_리포트.xlsx"

# 스크립트를 어느 위치에서 실행하든 backend/의 `app` 패키지를 import할 수 있게 경로를 추가한다.
sys.path.insert(0, str(BACKEND_DIR))

import openpyxl  # noqa: E402

from app.core.database import SessionLocal  # noqa: E402
from app.core.pricing import is_price_known  # noqa: E402
from app.models.usage_log import UsageLog  # noqa: E402
from app.services import usage_service  # noqa: E402

DAILY_HEADERS = [
    "날짜", "호출 수", "성공 수", "실패 수",
    "입력 토큰", "출력 토큰", "총 토큰", "총 비용(USD)", "평균 응답시간(ms)",
]
DETAIL_HEADERS = [
    "시각(UTC)", "기능", "모델", "단가 확인됨",
    "입력 토큰", "출력 토큰", "총 토큰", "도구 호출 수",
    "응답시간(ms)", "비용(USD)", "성공 여부", "오류 메시지",
]


def build_workbook(days: int) -> openpyxl.Workbook:
    db = SessionLocal()
    try:
        daily = usage_service.get_daily_usage(db, days=days)
        cutoff = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=days)
        detail_rows = (
            db.query(UsageLog)
            .filter(UsageLog.created_at >= cutoff)
            .order_by(UsageLog.created_at.desc())
            .all()
        )
    finally:
        db.close()

    wb = openpyxl.Workbook()

    summary_sheet = wb.active
    summary_sheet.title = "일별 요약"
    summary_sheet.append(DAILY_HEADERS)
    for d in daily:
        summary_sheet.append(
            [
                d.날짜, d.호출_수, d.성공_수, d.실패_수,
                d.입력_토큰, d.출력_토큰, d.총_토큰, d.총_비용_USD, d.평균_응답시간_ms,
            ]
        )

    detail_sheet = wb.create_sheet("호출 상세")
    detail_sheet.append(DETAIL_HEADERS)
    for row in detail_rows:
        detail_sheet.append(
            [
                row.created_at.isoformat(sep=" ", timespec="seconds"),
                row.feature,
                row.model,
                "Y" if is_price_known(row.model) else "N(단가표 갱신 필요)",
                row.input_tokens,
                row.output_tokens,
                row.total_tokens,
                row.tool_call_count,
                row.latency_ms,
                row.cost_usd,
                "성공" if row.success else "실패",
                row.error_message or "",
            ]
        )

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--days", type=int, default=30, help="집계 대상 기간(일). 기본 30일.")
    args = parser.parse_args()

    wb = build_workbook(args.days)
    wb.save(EXCEL_PATH)

    daily_count = wb["일별 요약"].max_row - 1
    detail_count = wb["호출 상세"].max_row - 1
    print(f"[완료] '{EXCEL_PATH.name}' 생성 - 일별 요약 {daily_count}행, 호출 상세 {detail_count}행")
    if detail_count == 0:
        print("[참고] 기록된 호출이 없습니다. AI상담(/chat)을 몇 번 호출한 뒤 다시 실행해보세요.")


if __name__ == "__main__":
    main()
